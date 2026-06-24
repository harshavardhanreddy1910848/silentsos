import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import os

def build_excel_report():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, "load_test_metrics.json")
    
    if not os.path.exists(json_path):
        print(f"Error: Raw metrics file not found at: {json_path}")
        return

    with open(json_path, "r") as f:
        data = json.load(f)

    summary = data["summary"]
    time_series = data["timeSeries"]

    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Styles
    # ----------------------------------------------------
    font_family = "Segoe UI"
    
    font_title = Font(name=font_family, size=18, bold=True, color="1F4E78")
    font_section = Font(name=font_family, size=13, bold=True, color="2C3E50")
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_body = Font(name=font_family, size=10)
    font_bold = Font(name=font_family, size=10, bold=True)
    
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_summary_hdr = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    fill_kpi = PatternFill(start_color="EAEDED", end_color="EAEDED", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # ----------------------------------------------------
    # Sheet 1: Dashboard
    # ----------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Load Test Summary"
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash['A1'] = "SilentSOS Distress Platform - Concurrency Load Test Executive Summary"
    ws_dash['A1'].font = font_title
    ws_dash.row_dimensions[1].height = 35
    
    # KPI Grid Slabs
    kpis = [
        ("Virtual Concurrent Users", f"{summary['concurrency']} VU", "Target mock capacity load"),
        ("Execution Duration", f"{summary['durationSeconds']} Seconds", "Continuous runtime duration"),
        ("Total Requests Sent", f"{summary['totalRequests']:,}", "Overall network transactions"),
        ("Requests Per Second (RPS)", f"{summary['rps']:.2f} req/sec", "Mean throughput benchmark"),
        ("Average Response Time", f"{summary['avgLatencyMs']:.2f} ms", "Latency mean average speed"),
        ("Min Response Time", f"{summary['minLatencyMs']} ms", "Fastest transaction speed"),
        ("Max Response Time", f"{summary['maxLatencyMs']} ms", "Slowest transaction peak"),
        ("Success Rate Status", f"{summary['successRate']:.2f}%", "Successful HTTP response code %)"),
    ]
    
    # Place KPIs in a nice two-column dashboard
    for idx, (label, val, desc) in enumerate(kpis):
        row = 4 + idx
        ws_dash.cell(row=row, column=1, value=label).font = font_bold
        ws_dash.cell(row=row, column=2, value=val).font = Font(name=font_family, size=11, bold=True, color="1F4E78")
        ws_dash.cell(row=row, column=3, value=desc).font = Font(name=font_family, size=9, italic=True)
        
        for col in range(1, 4):
            cell = ws_dash.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = fill_kpi
        ws_dash.row_dimensions[row].height = 24
        
    # Auto sizes
    ws_dash.column_dimensions['A'].width = 30
    ws_dash.column_dimensions['B'].width = 25
    ws_dash.column_dimensions['C'].width = 40

    # ----------------------------------------------------
    # Sheet 2: Time Series Logs
    # ----------------------------------------------------
    ws_logs = wb.create_sheet(title="Time Series Logs")
    ws_logs.views.sheetView[0].showGridLines = True
    
    headers = ["Second Elapsed", "Requests Executed", "Average Latency (ms)"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws_logs.cell(row=1, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws_logs.row_dimensions[1].height = 25
    ws_logs.freeze_panes = 'A2'
    
    for row_idx, item in enumerate(time_series, start=2):
        ws_logs.cell(row=row_idx, column=1, value=item["second"]).alignment = align_center
        ws_logs.cell(row=row_idx, column=2, value=item["requests"]).alignment = align_center
        ws_logs.cell(row=row_idx, column=3, value=item["avgLatencyMs"]).alignment = align_center
        
        for col in range(1, 4):
            cell = ws_logs.cell(row=row_idx, column=col)
            cell.font = font_body
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = fill_zebra
        ws_logs.row_dimensions[row_idx].height = 20
        
    ws_logs.column_dimensions['A'].width = 20
    ws_logs.column_dimensions['B'].width = 25
    ws_logs.column_dimensions['C'].width = 25
    
    # Save spreadsheet to workspace root
    root_path = os.path.dirname(script_dir)
    out_path = os.path.join(root_path, "load_test_results.xlsx")
    wb.save(out_path)
    print(f"[SUCCESS] Load test results Excel sheet created at: {out_path}")

if __name__ == "__main__":
    build_excel_report()
